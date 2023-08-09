# 导入os模块和openpyxl模块
import os
import openpyxl

# 定义两个文件夹的路径
imagespro_path = "F:\\DataBase\\dataset\\Uintfog\\imagespro"
xmlspro_path = "F:\\DataBase\\dataset\\Uintfog\\xmlspro"

# 定义一个excel文件的路径
name2_path = "F:\\DataBase\\dataset\\Uintfog\\name3.xlsx"

# 如果excel文件不存在，就创建一个新的工作簿，并添加一个工作表
if not os.path.exists(name2_path):
    wb = openpyxl.Workbook()
    ws = wb.active
else:
    # 如果excel文件存在，就打开它，并获取第一个工作表
    wb = openpyxl.load_workbook(name2_path)
    ws = wb.worksheets[0]

# 定义一个a排序法的函数，输入一个文件名列表，输出一个按照a排序法排序后的文件名列表
def a_sort(file_list):
    # 先按照文件名长度排序，短的在前面
    file_list.sort(key=len)
    # 再按照字典序排序，数字在前面，字母在后面
    # 但是要先比较文件名中的第一个数字，再比较第二个数字，以此类推，直到比较完所有的数字
    # 为了实现这个功能，可以使用lambda表达式来定义一个自定义的排序键
    # 这里使用了map函数和tuple函数来将文件名中的数字转换成元组，方便比较
    file_list.sort(key=lambda x: tuple(map(int, x.split("_"))))
    return file_list

# 获取imagespro文件夹下所有的文件名，并使用a排序法排序，并写入A列
row = 1 # 行号从1开始
file_list = os.listdir(imagespro_path) # 获取文件名列表
file_list = a_sort(file_list) # 使用a排序法排序
for file in file_list:
    ws.cell(row=row, column=1).value = file # 写入A列
    row += 1 # 行号加1

# 获取xmlspro文件夹下所有的文件名，并使用a排序法排序，并写入B列
row = 1 # 行号从1开始
file_list = os.listdir(xmlspro_path) # 获取文件名列表
file_list = a_sort(file_list) # 使用a排序法排序
for file in file_list:
    ws.cell(row=row, column=2).value = file # 写入B列
    row += 1 # 行号加1

# 保存并关闭excel文件
wb.save(name2_path)
wb.close()
