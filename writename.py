# 导入os模块和openpyxl模块
import os
import openpyxl

# 定义两个文件夹的路径
imagespro_path = "F:\\DataBase\\dataset\\Uintfog\\imagespro"
xmlspro_path = "F:\\DataBase\\dataset\\Uintfog\\xmlspro"

# 定义一个excel文件的路径
name_path = "F:\\DataBase\\dataset\\Uintfog\\name.xlsx"

# 如果excel文件不存在，就创建一个新的工作簿，并添加一个工作表
if not os.path.exists(name_path):
    wb = openpyxl.Workbook()
    ws = wb.active
else:
    # 如果excel文件存在，就打开它，并获取第一个工作表
    wb = openpyxl.load_workbook(name_path)
    ws = wb.worksheets[0]

# 获取imagespro文件夹下所有的文件名，并写入A列
row = 1 # 行号从1开始
for file in os.listdir(imagespro_path):
    ws.cell(row=row, column=1).value = file # 写入A列
    row += 1 # 行号加1

# 获取xmlspro文件夹下所有的文件名，并写入B列
row = 1 # 行号从1开始
for file in os.listdir(xmlspro_path):
    ws.cell(row=row, column=2).value = file # 写入B列
    row += 1 # 行号加1

# 保存并关闭excel文件
wb.save(name_path)
wb.close()
